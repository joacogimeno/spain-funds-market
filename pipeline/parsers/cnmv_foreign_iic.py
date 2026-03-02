"""Parse CNMV Estadisticas File 3, Cuadro 7.1-7.3 — foreign IICs sold in Spain."""
import os
import re
import openpyxl
from ..config import find_cnmv_file


def _build_quarters(year, q, n=5):
    """Generate the last n quarter labels ending at year-Qq."""
    result = []
    for _ in range(n):
        result.insert(0, f'{year}-Q{q}')
        q -= 1
        if q == 0:
            q, year = 4, year - 1
    return result


def parse_cnmv_foreign_iic(cnmv_dir):
    """Parse Cuadro 7.1 for foreign IIC summary data.

    Returns dict with:
      - quarters: list of 5 quarter labels (oldest to latest)
      - summary: number/accounts/volume for funds, societies, total (5 quarters)
      - countries: number of IICs by country of origin (5 quarters)
      - distribution_volume: IIC count and volume by size bracket (Cuadro 7.2)
      - distribution_accounts: IIC count and accounts by account bracket (Cuadro 7.3)
    """
    filepath, quarter_label = find_cnmv_file(cnmv_dir, '3')
    if filepath is None:
        print(f"  Warning: No CNMV File 3 found in {cnmv_dir}")
        return {}

    # Derive the 5-quarter window from the detected file
    m = re.search(r'(\d{4})_(\d)T', os.path.basename(filepath))
    if m:
        quarters = _build_quarters(int(m.group(1)), int(m.group(2)))
    else:
        quarters = ['2024-Q3', '2024-Q4', '2025-Q1', '2025-Q2', '2025-Q3']

    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        'quarters': quarters,
        'summary': _parse_cuadro_71(wb),
        'distribution_volume': _parse_cuadro_72(wb),
        'distribution_accounts': _parse_cuadro_73(wb),
    }

    wb.close()
    return result


# Column mapping: columns B-F in the spreadsheet correspond to the 5 quarters
QUARTER_COLS = [2, 3, 4, 5, 6]  # B through F
PCT_COLS = {'qoq': 8, 'yoy': 9, 'ytd': 10}  # H, I, J


def _parse_cuadro_71(wb):
    """Parse Cuadro 7.1 - summary and country breakdown."""
    ws = wb['Cuadro 7.1']
    rows = list(ws.iter_rows(min_row=1, values_only=False))

    def _row_vals(r):
        return {c.column: c.value for c in r}

    # Build a lookup by row number
    data = {}
    for r in rows:
        rv = _row_vals(r)
        row_num = r[0].row
        data[row_num] = rv

    summary = {}

    # Number of IICs (rows 8-10)
    summary['num_iic'] = {
        'fondos': _extract_series(data.get(8, {})),
        'sociedades': _extract_series(data.get(9, {})),
        'total': _extract_series(data.get(10, {})),
    }

    # Investor accounts (rows 13-15)
    summary['accounts'] = {
        'fondos': _extract_series(data.get(13, {})),
        'sociedades': _extract_series(data.get(14, {})),
        'total': _extract_series(data.get(15, {})),
    }

    # Investment volume in thousands (rows 18-20)
    summary['volume_k'] = {
        'fondos': _extract_series(data.get(18, {})),
        'sociedades': _extract_series(data.get(19, {})),
        'total': _extract_series(data.get(20, {})),
    }

    # Countries (rows 23-37)
    countries = []
    for row_num in range(23, 38):
        rv = data.get(row_num, {})
        label = str(rv.get(1, '') or '').strip()
        if not label or label.startswith(('1.', '2.', '3.')) is False:
            # Extract country name after "N. "
            pass
        if label and any(c.isalpha() for c in label):
            # Remove leading number prefix like "1. ", "15: "
            name = label
            for sep in ['. ', ': ']:
                if sep in name:
                    name = name.split(sep, 1)[1]
            series = _extract_series(rv)
            if series['values']:
                countries.append({
                    'name': name,
                    **series,
                })

    summary['countries'] = countries
    return summary


def _parse_cuadro_72(wb):
    """Parse Cuadro 7.2 - distribution by investment volume."""
    ws = wb['Cuadro 7.2']
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    data = {}
    for r in rows:
        data[r[0].row] = {c.column: c.value for c in r}

    # Number of IICs by bracket (rows 8-17)
    brackets_count = []
    for row_num in range(8, 17):
        rv = data.get(row_num, {})
        label = str(rv.get(1, '') or '').strip()
        if label and label != 'TOTAL':
            brackets_count.append({'bracket': label, **_extract_series(rv)})

    # Total
    total_count = _extract_series(data.get(17, {}))

    # Volume by bracket (rows 20-30)
    brackets_volume = []
    for row_num in range(20, 30):
        rv = data.get(row_num, {})
        label = str(rv.get(1, '') or '').strip()
        if label and label != 'TOTAL':
            brackets_volume.append({'bracket': label, **_extract_series(rv)})

    return {
        'by_count': brackets_count,
        'total_count': total_count,
        'by_volume': brackets_volume,
    }


def _parse_cuadro_73(wb):
    """Parse Cuadro 7.3 - distribution by account count."""
    ws = wb['Cuadro 7.3']
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    data = {}
    for r in rows:
        data[r[0].row] = {c.column: c.value for c in r}

    brackets_count = []
    for row_num in range(8, 17):
        rv = data.get(row_num, {})
        label = str(rv.get(1, '') or '').strip()
        if label and label != 'TOTAL':
            brackets_count.append({'bracket': label, **_extract_series(rv)})

    total_count = _extract_series(data.get(17, {}))

    brackets_accounts = []
    for row_num in range(20, 30):
        rv = data.get(row_num, {})
        label = str(rv.get(1, '') or '').strip()
        if label and label != 'TOTAL':
            brackets_accounts.append({'bracket': label, **_extract_series(rv)})

    return {
        'by_count': brackets_count,
        'total_count': total_count,
        'by_accounts': brackets_accounts,
    }


def _extract_series(row_vals):
    """Extract quarterly values and percentage changes from a row."""
    values = []
    for col in QUARTER_COLS:
        v = _num(row_vals.get(col))
        values.append(v)

    pct = {}
    for key, col in PCT_COLS.items():
        pct[key] = _num(row_vals.get(col))

    return {'values': values, 'pct': pct}


def _num(val):
    """Convert to float, returning None if not numeric."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ('', '-'):
            return None
        try:
            return float(val.replace(',', '.'))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return float(val)
    return None
