"""Build inversis.json — comprehensive Inversis snapshot.

Covers two distinct business lines:
  1. Banco Inversis (depositary) — custody relationships, fees, market share
  2. Inversis Gestión (gestora) — AUM, growth, fund breakdown

Also includes:
  - SICAV market context and Inversis Gestión's SICAV book
  - Non-captive opportunity pipeline
  - Competitive positioning radar
"""
import csv
import glob
import json
import os

import openpyxl

from ..config import DATA_DIR, OUTPUT_DIR, find_cnmv_file, get_snapshot_folders
from ..parsers.cnmv_fees import parse_cnmv_fees
from ..parsers.ranking_grupos import parse_ranking_grupos

CNMV_DIR = os.path.join(DATA_DIR, 'CNMV Estadisticas')
INSIGHTS_XL = os.path.join(CNMV_DIR, 'inversis_depositary_insights_2025Q3.xlsx')
CLIENT_CSV   = os.path.join(CNMV_DIR, 'inversis_client_fees_2025Q3.csv')

INVERSIS_DEPOSITARIO = 'BANCO INVERSIS, S.A.'
INVERSIS_GESTORA     = 'INVERSIS GESTIÓN, S.A., SGIIC'

# Captive groupings: gestora group → their own depositary (excluded from opportunity)
CAPTIVE_MAP = {
    'CAIXABANK': 'CECABANK',
    'BBVA': 'BBVA',
    'SANTANDER': 'CACEIS',
    'BANKINTER': 'BANKINTER',
    'KUTXABANK': 'CECABANK',
    'SABADELL': 'CECABANK',
    'IBERCAJA': 'CECABANK',
    'UNICAJA': 'CECABANK',
    'MAPFRE': 'BNP',
    'CREDIT AGRICOLE': 'BNP',
}


def _sf(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _load_insights_xl():
    """Load the Inversis insights Excel file into dict of row-lists."""
    if not os.path.exists(INSIGHTS_XL):
        return {}
    wb = openpyxl.load_workbook(INSIGHTS_XL, read_only=True)
    result = {}
    for sh in ['inversis_by_gestora', 'inversis_by_group', 'depositary_ranking', 'inversis_isin_detail']:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        result[sh] = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
    wb.close()
    return result


def _load_client_csv():
    """Load inversis_client_fees_2025Q3.csv."""
    if not os.path.exists(CLIENT_CSV):
        return []
    with open(CLIENT_CSV, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)


def _parse_sicav(latest_folder):
    """Parse the latest SICAV monthly file. Returns (market_summary, group_ranking)."""
    datos_dir = os.path.join(latest_folder, 'Datos Generales')
    sicav_files = glob.glob(os.path.join(datos_dir, '*-SICAV.xlsx'))
    if not sicav_files:
        return {}, []

    wb = openpyxl.load_workbook(sicav_files[0], read_only=True)
    if 'Sociedades' not in wb.sheetnames:
        wb.close()
        return {}, []

    ws = wb['Sociedades']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row (contains 'Rkg.' or 'PATRIMONIO')
    header_idx = None
    for i, r in enumerate(rows):
        vals = [v for v in r if v is not None]
        if any('Rkg' in str(v) or 'PATRIMONIO' in str(v).upper() for v in vals):
            header_idx = i
            break

    if header_idx is None:
        return {}, []

    # Parse data rows: [rank, group, gestora, patrimonio, count, shareholders]
    # Some rows are sub-gestoras (no rank number) - we skip those for group ranking
    groups = []
    total_aum_k = 0.0
    total_count = 0
    total_shareholders = 0
    inversis_aum_k = 0.0
    inversis_count = 0

    for r in rows[header_idx + 1:]:
        vals = [v for v in r if v is not None]
        if not vals:
            continue

        # Detect total row ("T O T A L" with spaces in Spanish files)
        first_str = str(vals[0]).upper().replace(' ', '')
        if 'TOTAL' in first_str:
            # Extract totals
            nums = [v for v in vals if isinstance(v, (int, float))]
            if nums:
                total_aum_k = _sf(nums[0])
            if len(nums) > 1:
                total_count = int(nums[1])
            if len(nums) > 2:
                total_shareholders = int(nums[2])
            break

        # Only process rows with a rank number (group-level, not sub-gestora)
        if not isinstance(vals[0], (int, float)):
            # Sub-gestora row: check if it's Inversis gestora
            name = str(vals[0]).strip().upper()
            if 'INVERSIS' in name:
                nums = [v for v in vals if isinstance(v, (int, float))]
                if nums:
                    inversis_aum_k = _sf(nums[0])
                if len(nums) > 1:
                    inversis_count = int(nums[1])
            continue

        rank = int(vals[0])
        group = str(vals[1]).strip() if len(vals) > 1 else ''
        # gestora might be in col 2 or missing (standalone SICAV groups)
        gestora = ''
        patrimonio = 0.0
        count = 0
        shareholders = 0

        nums = [v for v in vals[1:] if isinstance(v, (int, float))]
        strs = [v for v in vals[1:] if isinstance(v, str)]

        group = strs[0].strip() if strs else ''
        gestora = strs[1].strip() if len(strs) > 1 else ''
        if nums:
            patrimonio = _sf(nums[0])
        if len(nums) > 1:
            count = int(nums[1])
        if len(nums) > 2:
            shareholders = int(nums[2])

        groups.append({
            'rank':         rank,
            'group':        group,
            'gestora':      gestora,
            'aum_m':        round(patrimonio / 1000, 1),
            'count':        count,
            'shareholders': shareholders,
        })

    return {
        'total_market_aum_bn':   round(total_aum_k / 1_000_000, 2),
        'total_market_count':    total_count,
        'total_shareholders':    total_shareholders,
        'inversis_gestora_aum_m': round(inversis_aum_k / 1000, 1),
        'inversis_gestora_count': inversis_count,
    }, groups[:30]


def _build_gestora_series(snapshots):
    """Build monthly AUM time series for Inversis Gestión from ranking_grupos.

    Inversis Gestión appears as a sub-gestora inside the 'BANCA MARCH' group.
    """
    series = []
    for label, folder in snapshots:
        groups = parse_ranking_grupos(folder)
        found = False
        for g in groups:
            # Check top-level group name
            name = g.get('name', '').upper()
            # First check sub-gestoras list
            for sg in g.get('gestoras', []):
                sg_name = sg.get('name', '').upper()
                if 'INVERSIS' in sg_name:
                    pat = sg.get('patrimonio')
                    if pat and pat > 0:
                        series.append({'label': label, 'aum_bn': round(pat / 1_000_000, 3)})
                        found = True
                    break
            if found:
                break
            # Fallback: top-level group
            if 'INVERSIS' in name and 'BANCA MARCH' not in name:
                pat = g.get('patrimonio')
                if pat and pat > 0:
                    series.append({'label': label, 'aum_bn': round(pat / 1_000_000, 3)})
                    found = True
                break
    return series


def build_inversis():
    """Build Inversis comprehensive snapshot JSON."""
    snapshots = get_snapshot_folders()
    if not snapshots:
        return {}

    _, latest_folder = snapshots[-1]
    _, cnmv_date = find_cnmv_file(CNMV_DIR, 'Anexo')
    cnmv_date = cnmv_date or 'unknown'

    # ── 1. Load Inversis insights Excel ─────────────────────────────
    xl = _load_insights_xl()
    csv_rows = _load_client_csv()

    # ── 2. Depositary book: by gestora ─────────────────────────────
    by_gestora_xl = xl.get('inversis_by_gestora', [])
    by_gestora = []
    total_aum_k  = 0.0
    total_rev_k  = 0.0

    for r in by_gestora_xl:
        aum_k   = _sf(r.get('aum_k'))
        rev_k   = _sf(r.get('est_deposit_fee_rev_k'))
        # aum_w_deposit_fee_pct is a fraction (e.g. 0.000325 = 3.25 bps)
        # convert: pct * 10000 → bps only when pct < 0.01 (already in fractional form)
        raw_fee = _sf(r.get('aum_w_deposit_fee_pct'))  # e.g. 0.0325% → 3.25 bps
        wtd_bps = round(raw_fee * 100, 2)  # 0.0325 * 100 = 3.25 bps
        gestora = str(r.get('Sociedad Gestora', '')).strip()
        short   = gestora.split(',')[0].split('SGIIC')[0].strip()
        total_aum_k += aum_k
        total_rev_k += rev_k
        by_gestora.append({
            'gestora':       gestora,
            'gestora_short': short,
            'classes':       int(_sf(r.get('classes', 0))),
            'funds':         int(_sf(r.get('funds', 0))),
            'aum_m':         round(aum_k / 1000, 1),
            'wtd_fee_bps':   wtd_bps,
            'est_rev_k':     round(rev_k, 1),
            'investors':     int(_sf(r.get('investors', 0))),
        })

    by_gestora.sort(key=lambda x: -x['aum_m'])

    # ── 3. Depositary book: by group ────────────────────────────────
    by_group_xl = xl.get('inversis_by_group', [])
    by_group = []
    for r in by_group_xl:
        aum_k = _sf(r.get('aum_k'))
        by_group.append({
            'group':    str(r.get('Grupo Financiero', '')).strip(),
            'gestoras': int(_sf(r.get('gestoras', 0))),
            'classes':  int(_sf(r.get('classes', 0))),
            'funds':    int(_sf(r.get('funds', 0))),
            'aum_m':    round(aum_k / 1000, 1),
            'est_rev_k': round(_sf(r.get('est_deposit_fee_rev_k')), 1),
        })
    by_group.sort(key=lambda x: -x['aum_m'])

    # ── 4. Depositary market ranking ───────────────────────────────
    dep_ranking_xl = xl.get('depositary_ranking', [])
    all_dep_aum    = sum(_sf(r.get('aum_k')) for r in dep_ranking_xl)
    market_ranking = []
    inv_rank_aum   = None
    inv_rank_g     = None

    ranked_by_aum = sorted(dep_ranking_xl, key=lambda r: -_sf(r.get('aum_k')))
    # Sort by gestora_count for rank by gestoras
    ranked_by_gest = sorted(dep_ranking_xl, key=lambda r: -_sf(r.get('classes', 0)))

    for i, r in enumerate(ranked_by_aum):
        dep   = str(r.get('Entidad Depositaria', '')).strip()
        aum_k = _sf(r.get('aum_k'))
        is_inv = INVERSIS_DEPOSITARIO in dep
        if is_inv:
            inv_rank_aum = i + 1
        market_ranking.append({
            'rank':             i + 1,
            'depositario':      dep,
            'depositario_short': dep.split(',')[0].split('S.A.')[0].strip(),
            'aum_bn':           round(aum_k / 1_000_000, 2),
            'market_share_pct': round(aum_k / all_dep_aum * 100, 1) if all_dep_aum else 0,
            'fund_count':       int(_sf(r.get('funds', 0))),
            'class_count':      int(_sf(r.get('classes', 0))),
            'is_inversis':      is_inv,
        })

    # Rank by gestoras served
    for i, r in enumerate(ranked_by_gest):
        dep = str(r.get('Entidad Depositaria', '')).strip()
        if INVERSIS_DEPOSITARIO in dep:
            inv_rank_g = i + 1
            break

    # ── 5. Depositary summary metrics ──────────────────────────────
    inv_entry = next(
        (r for r in dep_ranking_xl if INVERSIS_DEPOSITARIO in str(r.get('Entidad Depositaria', ''))),
        None,
    )
    inv_aum_k      = _sf(inv_entry.get('aum_k')) if inv_entry else total_aum_k
    inv_class_cnt  = int(_sf(inv_entry.get('classes'))) if inv_entry else len(by_gestora)
    inv_fund_cnt   = int(_sf(inv_entry.get('funds'))) if inv_entry else 0
    inv_share_pct  = round(inv_aum_k / all_dep_aum * 100, 1) if all_dep_aum else 0

    # Weighted fee across book (from CSV)
    csv_num = sum(_sf(r.get('aum_bn_eur')) * _sf(r.get('weighted_deposit_fee_bps'))
                  for r in csv_rows)
    csv_denom = sum(_sf(r.get('aum_bn_eur')) for r in csv_rows)
    wtd_fee_bps = round(csv_num / csv_denom, 2) if csv_denom else 0.0
    est_rev_m   = round(sum(_sf(r.get('estimated_annual_deposit_fee_rev_m_eur')) for r in csv_rows), 2)

    # ── 6. Market fee context — load from cnmv_depositaria.json ────
    # (cnmv_depositaria is always generated before inversis in the pipeline)
    dep_json_path = os.path.join(OUTPUT_DIR, 'cnmv_depositaria.json')
    dep_json = {}
    if os.path.exists(dep_json_path):
        with open(dep_json_path, encoding='utf-8') as f:
            dep_json = json.load(f)

    dep_summary = dep_json.get('summary', {})
    fee_by_dep  = dep_json.get('fee_by_depositario', [])

    # Market fee stats from fee_by_depositario (avg_fee is in % → * 100 for bps)
    all_avg_bps = sorted(_sf(d.get('avg_fee')) * 100 for d in fee_by_dep if d.get('avg_fee'))
    n = len(all_avg_bps)
    p25 = all_avg_bps[n // 4]    if n > 0 else 0
    p50 = all_avg_bps[n // 2]    if n > 0 else 0
    p75 = all_avg_bps[3*n // 4]  if n > 0 else 0

    # Inversis avg fee from fee_by_depositario
    inv_fee_entry = next((d for d in fee_by_dep if d.get('is_inversis')), None)
    inv_avg_fee_bps = round(_sf(inv_fee_entry.get('avg_fee', 0)) * 100, 1) if inv_fee_entry else 0

    # ── 7. Opportunity targets — reuse from cnmv_depositaria.json ──
    raw_targets  = dep_json.get('opportunity_targets', [])
    non_captive  = [t for t in raw_targets if not t.get('is_captive')]
    targets = []
    for t in raw_targets[:40]:
        short     = t.get('gestora', '').split(',')[0].split('SGIIC')[0].strip()
        dep_short = t.get('current_depositario', '').split(',')[0].split('S.A.')[0].strip()
        targets.append({
            'gestora':             t.get('gestora', ''),
            'gestora_short':       short,
            'current_depositario': dep_short,
            'grupo':               t.get('grupo', ''),
            'aum_m':               t.get('total_aum_m', 0),
            'fund_count':          t.get('fund_count', 0),
            'avg_fee_bps':         round(_sf(t.get('avg_depo_fee', 0)) * 100, 2),
            'potential_rev_k':     t.get('potential_revenue_k', 0),
            'is_captive':          t.get('is_captive', False),
        })

    nc_targets      = [t for t in targets if not t['is_captive']]
    addressable_aum = _sf(dep_summary.get('addressable_aum_bn', 0)) or round(
        sum(t['aum_m'] for t in nc_targets) / 1000, 1)
    potential_rev_m = _sf(dep_summary.get('potential_revenue_m', 0)) or round(
        sum(t['potential_rev_k'] for t in nc_targets) / 1000, 1)

    # ── 8. Inversis Gestión (gestora metrics) ──────────────────────
    # Inversis Gestión appears as a sub-gestora under the BANCA MARCH group
    latest_groups = parse_ranking_grupos(latest_folder)
    inv_gest_data = None
    for g in latest_groups:
        for sg in g.get('gestoras', []):
            if 'INVERSIS' in sg.get('name', '').upper():
                inv_gest_data = sg
                break
        if inv_gest_data:
            break
    if not inv_gest_data:
        # Fallback: top-level group
        for g in latest_groups:
            name = g.get('name', '').upper()
            if 'INVERSIS' in name and 'BANCA MARCH' not in name:
                inv_gest_data = g
                break

    # From CSV (Inversis Gestión row)
    inv_csv = next((r for r in csv_rows if 'INVERSIS' in str(r.get('Sociedad Gestora', '')).upper()
                    and 'BANCA MARCH' not in str(r.get('Sociedad Gestora', '')).upper()), None)

    gestora_section = {
        'name':       INVERSIS_GESTORA,
        'aum_bn':     round(inv_gest_data['patrimonio'] / 1_000_000, 3) if inv_gest_data else (
                          _sf(inv_csv.get('aum_bn_eur')) if inv_csv else None),
        'growth_1y':  inv_gest_data.get('var_1y') if inv_gest_data else None,
        'growth_ytd': inv_gest_data.get('var_ytd') if inv_gest_data else None,
        'fund_count': int(_sf(inv_csv.get('funds'))) if inv_csv else (
                          int(inv_gest_data.get('num_isin', 0)) if inv_gest_data else 0),
        'class_count': int(_sf(inv_csv.get('classes'))) if inv_csv else 0,
        'fee_bps_paid': round(_sf(inv_csv.get('weighted_deposit_fee_bps')) if inv_csv else 0, 2),
        'est_rev_k':    round(_sf(inv_csv.get('estimated_annual_deposit_fee_rev_m_eur')) * 1000
                              if inv_csv else 0, 1),
        'aum_series':   _build_gestora_series(snapshots),
    }

    # ── 9. SICAV section ───────────────────────────────────────────
    sicav_summary, sicav_ranking = _parse_sicav(latest_folder)

    # ── 10. Positioning radar (reuse from insights.py values) ──────
    radar = [
        {'metric': 'AUM Scale',    'inversis': 25, 'market_avg': 50, 'top_peer': 90},
        {'metric': 'Growth Rate',  'inversis': 35, 'market_avg': 50, 'top_peer': 85},
        {'metric': 'Fund Diversity','inversis': 60, 'market_avg': 50, 'top_peer': 80},
        {'metric': 'Client Reach', 'inversis': 40, 'market_avg': 50, 'top_peer': 95},
        {'metric': 'Technology',   'inversis': 75, 'market_avg': 50, 'top_peer': 70},
        {'metric': 'Independence', 'inversis': 80, 'market_avg': 40, 'top_peer': 90},
    ]

    # ── 11. Assemble output ────────────────────────────────────────
    return {
        'date': cnmv_date,
        'depositary': {
            'aum_bn':           round(inv_aum_k / 1_000_000, 2),
            'class_count':      inv_class_cnt,
            'fund_count':       inv_fund_cnt,
            'gestora_count':    len(by_gestora),
            'market_share_pct': inv_share_pct,
            'rank_aum':         inv_rank_aum,
            'rank_gestoras':    inv_rank_g,
            'wtd_fee_bps':      wtd_fee_bps,
            'est_annual_rev_m': est_rev_m,
            'by_gestora':       by_gestora,
            'by_group':         by_group,
            'market_ranking':   market_ranking,
        },
        'gestora': gestora_section,
        'sicav': {
            **sicav_summary,
            'market_ranking': sicav_ranking,
        },
        'market_context': {
            'total_depositario_aum_bn': round(all_dep_aum / 1_000_000, 1),
            'total_depositarios':       len(fee_by_dep),
            'inv_avg_fee_bps':          inv_avg_fee_bps,
            'mkt_dep_fee_p25_bps':      round(p25, 2),
            'mkt_dep_fee_p50_bps':      round(p50, 2),
            'mkt_dep_fee_p75_bps':      round(p75, 2),
        },
        'opportunity': {
            'addressable_aum_bn':    round(addressable_aum, 1),
            'target_gestoras_count': len(non_captive),
            'potential_revenue_m':   potential_rev_m,
            'targets':               nc_targets[:40],
        },
        'positioning': {
            'radar': radar,
        },
    }
