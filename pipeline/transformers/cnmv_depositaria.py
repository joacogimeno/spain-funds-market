"""Transform CNMV registry + fee data into cnmv_depositaria.json for the Depositaría dashboard.

Also reads inversis_depositary_insights_2025Q3.xlsx for:
- Full-market depositary AUM ranking (depositary_ranking sheet)
- Inversis client detail with estimated deposit fee revenue (by_gestora sheet)
- Computed qualitative analysis for the Depositar\u00eda tab

Pensiones / EPSV and full-scope Alternativos AUM are NOT in CNMV public feeds.
We overlay published values from Data/published/depositary_published_2025Q4.json
(FundsPeople IX 2026 ranking + Cecabank press releases + INVERCO totals).
"""
import json
import os
import openpyxl

from ..parsers.cnmv_registry import parse_cnmv_registry
from ..parsers.cnmv_fees import parse_cnmv_fees, VOCACION_MAP
from ..parsers.cnmv_sicav import parse_cnmv_sicav
from ..parsers.cnmv_alternativos import parse_cnmv_alternativos
from ..config import DATA_DIR, find_cnmv_file

INSIGHTS_FILE = 'inversis_depositary_insights_2025Q3.xlsx'
PUBLISHED_FILE = os.path.join(DATA_DIR, 'published', 'depositary_published_2025Q4.json')


def _load_published_overlay():
    """Load FundsPeople-published overlay (pensions, EPSV, true alts AUM)."""
    if not os.path.exists(PUBLISHED_FILE):
        return None
    with open(PUBLISHED_FILE, encoding='utf-8') as f:
        return json.load(f)


def _match_published(depositario_name, published):
    """Return the published row whose match-key is a substring of the depositario name."""
    if not published:
        return None
    name_up = (depositario_name or '').upper()
    for row in published.get('depositaries', []):
        key = (row.get('depositario_match') or '').upper()
        if key and key in name_up:
            return row
    return None


# Known captive pairs: grupo keyword → depositario keyword
# These are banks where the gestora uses the parent bank's depositaría by policy
CAPTIVE_KEYWORDS = [
    ('BBVA', 'BBVA'),
    ('BILBAO VIZCAYA', 'BBVA'),
    ('BILBAO VIZCAYA', 'BILBAO VIZCAYA'),
    ('LA CAIXA', 'CECABANK'),  # CaixaBank (grupo="LA CAIXA") uses Cecabank
    ('CAIXABANK', 'CECABANK'),
    ('SANTANDER', 'CACEIS'),  # Santander sold custody to CACEIS
    ('BANKINTER', 'BANKINTER'),
    ('CREDIT AGRICOLE', 'BNP PARIBAS'),  # Sabadell grupo=Credit Agricole, uses BNP
    ('CREDIT AGRICOLE', 'CACEIS'),
    ('KUTXABANK', 'CECABANK'),
    ('IBERCAJA', 'CECABANK'),
    ('UNICAJA', 'CECABANK'),
    ('ABANCA', 'CECABANK'),
    ('LABORAL KUTXA', 'CECABANK'),
    ('CAJA LABORAL', 'CAJA LABORAL'),  # Caja Laboral uses own depositaría
    ('CAJA RURAL', 'CECABANK'),
    ('BANCO COOPERATIVO', 'CECABANK'),  # Cajas rurales group
    ('BANCO COOPERATIVO', 'COOPERATIVO'),  # Gescooperativo uses own group depositaría
    ('MAPFRE', 'MAPFRE'),
    ('MEDIOLANUM', 'MEDIOLANUM'),
    ('DEUTSCHE BANK', 'DEUTSCHE'),
    ('RENTA 4', 'RENTA 4'),
    ('CAJA DE CREDITO DE LOS INGENIEROS', 'ENGINYERS'),  # Caja Ingenieros uses own depo
    ('SINGULAR BANK', 'SINGULAR'),  # Singular Bank uses own depo
]

INVERSIS_NAME = 'BANCO INVERSIS, S.A.'


def _is_captive(grupo, depositario):
    """Check if a gestora-depositario relationship is captive (parent bank uses own depositaría)."""
    grupo_up = (grupo or '').upper()
    depo_up = (depositario or '').upper()
    for g_kw, d_kw in CAPTIVE_KEYWORDS:
        if g_kw in grupo_up and d_kw in depo_up:
            return True
    return False


def _median(values):
    if not values:
        return 0
    s = sorted(values)
    n = len(s)
    if n % 2 == 1:
        return round(s[n // 2], 4)
    return round((s[n // 2 - 1] + s[n // 2]) / 2, 4)


def _read_sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    return [{header[i]: row[i] for i in range(min(len(header), len(row)))} for row in rows[1:]]


def _sf(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _load_insights(cnmv_dir):
    """Load inversis_depositary_insights_2025Q3.xlsx — returns dict of DataFrames-as-records."""
    fp = os.path.join(cnmv_dir, INSIGHTS_FILE)
    if not os.path.exists(fp):
        return {}
    wb = openpyxl.load_workbook(fp, data_only=True)
    result = {}
    for sheet in ['inversis_isin_detail', 'inversis_by_gestora', 'inversis_by_group',
                  'inversis_by_fund', 'depositary_ranking']:
        if sheet in wb.sheetnames:
            result[sheet] = _read_sheet_records(wb[sheet])
    wb.close()
    return result


def _build_qualitative_analysis(insights, inversis_stats, depositario_stats, opportunity_targets):
    """Generate computed qualitative insights for the Depositar\u00eda tab."""
    gest_records = insights.get('inversis_by_gestora', [])
    grp_records  = insights.get('inversis_by_group', [])

    total_rev_k = sum(_sf(r.get('est_deposit_fee_rev_k')) for r in gest_records)
    total_aum_k = _sf(inversis_stats.get('total_aum_bn', 0)) * 1_000_000
    book_eff_rate_bps = (total_rev_k / total_aum_k * 100 * 100) if total_aum_k > 0 else 0

    # Top 2 revenue clients
    top_rev = sorted(gest_records, key=lambda r: -_sf(r.get('est_deposit_fee_rev_k')))[:2]

    # Non-captive addressable targets
    nc_targets = [t for t in opportunity_targets if not t.get('is_captive')]
    addressable_aum_bn = sum(t.get('total_aum_m', 0) for t in nc_targets) / 1000

    # Rank by AUM in full market
    rank_records = insights.get('depositary_ranking', [])
    total_market_aum_k = sum(_sf(r.get('aum_k')) for r in rank_records)
    inv_rank = next((i + 1 for i, r in enumerate(
        sorted(rank_records, key=lambda x: -_sf(x.get('aum_k'))))
        if 'INVERSIS' in str(r.get('Entidad Depositaria', ''))), 0)

    insights_out = [
        {
            'type': 'market_position',
            'severity': 'info',
            'title': f'Inversis is #{inv_rank} depositary in Spain with {inversis_stats.get("inversis_market_share_pct", 0):.1f}% market share',
            'body': (
                f'Inversis custodies \u20AC{inversis_stats.get("inversis_aum_bn", 0):.1f}B across '
                f'{inversis_stats.get("inversis_gestora_count", 0)} gestoras and '
                f'{inversis_stats.get("inversis_fund_count", 0)} funds (FI + SICAV). '
                f'The overall depositary market (FI + SICAV, ex-foreign IICs / FIL / FCR / pensions) '
                f'totals \u20AC{total_market_aum_k/1e6:.0f}B, dominated by Cecabank (44.7%) and CACEIS (19.5%). '
                f'Inversis\u2019s position is structurally differentiated: it serves independent '
                f'boutiques and Andorran-origin private banking gestoras that are excluded from '
                f'captive bank depositaries.'
            ),
        },
        {
            'type': 'revenue',
            'severity': 'info',
            'title': f'Estimated annual depositary fee revenue: \u20AC{total_rev_k/1000:.1f}M at {book_eff_rate_bps:.1f} bps effective rate',
            'body': (
                f'The Inversis depositary book generates an estimated \u20AC{total_rev_k/1000:.2f}M/yr in depositary fees. '
                f'The top two fee contributors are '
                f'{str(top_rev[0].get("Sociedad Gestora","")).split(",")[0] if top_rev else ""} '
                f'(\u20AC{_sf(top_rev[0].get("est_deposit_fee_rev_k"))/1000:.2f}M) and '
                f'{str(top_rev[1].get("Sociedad Gestora","")).split(",")[0] if len(top_rev) > 1 else ""} '
                f'(\u20AC{_sf(top_rev[1].get("est_deposit_fee_rev_k"))/1000:.2f}M), '
                f'together representing '
                f'{(_sf(top_rev[0].get("est_deposit_fee_rev_k")) + _sf(top_rev[1].get("est_deposit_fee_rev_k"))) / total_rev_k * 100:.0f}% '
                f'of total depositary income.'
            ),
        },
        {
            'type': 'concentration_risk',
            'severity': 'medium',
            'title': 'Client concentration: top 2 groups account for 51% of AUM and revenue',
            'body': (
                'Andbank group (\u20AC3.23B, \u20AC2.12M fees) and Banca March group (\u20AC3.77B, \u20AC1.73M fees) '
                'together represent 57% of AUM under custody. The Banca March relationship will '
                'transition to arm\'s-length terms upon Euroclear acquisition (August 2026), '
                'creating both a repricing opportunity and an account retention priority. '
                'Andbank, as the highest-value third-party client, should be treated as a '
                'Tier 1 strategic account.'
            ),
        },
        {
            'type': 'business_development',
            'severity': 'opportunity',
            'title': f'\u20AC{addressable_aum_bn:.0f}B addressable market outside current Inversis relationships',
            'body': (
                f'{len(nc_targets)} non-captive gestoras totalling \u20AC{addressable_aum_bn:.0f}B AUM '
                f'currently use other depositaries. The most valuable non-captive targets include '
                f'{", ".join(str(t.get("gestora","")).split(",")[0] for t in nc_targets[:3])}. '
                f'Inversis\u2019s competitive advantage in this segment is its combination of '
                f'regulatory independence, operational breadth ('
                f'{inversis_stats.get("inversis_fund_count", 0)} funds across '
                f'{inversis_stats.get("inversis_gestora_count", 0)} gestoras), '
                f'and — post-Euroclear — tier-1 group backing without conflicting distribution interests.'
            ),
        },
    ]
    return insights_out


def build_cnmv_depositaria():
    """Build depositaría market data from CNMV Anexo A1.1 + A2.2."""
    cnmv_dir = os.path.join(DATA_DIR, 'CNMV Estadisticas')
    _, cnmv_date = find_cnmv_file(cnmv_dir, 'Anexo')
    cnmv_date = cnmv_date or 'unknown'

    # Parse CNMV Anexo A1.1 + A2.2 (FI registry + fees)
    registry = parse_cnmv_registry(cnmv_dir)
    fees = parse_cnmv_fees(cnmv_dir)

    # Parse SICAV registry + patrimonio from CNMV quarterly XML (SOCREGISTRO + SOCTRIM).
    # FI Anexo excludes SICAVs entirely, which understates depositary AuM by the
    # full SICAV book (~€16B market in 2025; ~€3B at Inversis).
    sicav_records, sicav_period = parse_cnmv_sicav(cnmv_dir)

    # Scrape CNMV per-entity pages for alternativo categories (FIL, FCR/SCR, SICC/FICC,
    # FILPE, FESE). No bulk XML exists for these; AUM is not exposed on the public HTML
    # pages so we track entity counts only and combine with the aggregate FIL AUM
    # (Cuadro 6.1: IICIL €7.16B + IICIICIL €0.83B = €8.0B Q2 2025).
    alt_cache = os.path.join(cnmv_dir, 'cnmv_alternativos_cache.json')
    alternativos_records = parse_cnmv_alternativos(alt_cache, verbose=False)

    # Load Inversis depositary insights (richer data for market ranking + revenue)
    insights = _load_insights(cnmv_dir)

    if not registry:
        return {}

    # Build ISIN → fee/AUM lookup from A2.2
    fee_by_isin = {}
    for r in fees:
        fee_by_isin[r['isin']] = {
            'depositary_fee': r['depositary_fee'],
            'patrimonio_k': r['patrimonio_k'],
            'mgmt_fee_aum': r['mgmt_fee_aum'],
            'ter': r['ter'],
            'category': r['category'],
            'vocacion': r['vocacion'],
            'investors': r['investors'],
        }

    # Join registry + fees by ISIN
    joined = []
    for r in registry:
        fee_data = fee_by_isin.get(r['isin'], {})
        joined.append({
            'fund_name': r['fund_name'],
            'isin': r['isin'],
            'share_class': r['share_class'],
            'gestora': r['gestora'],
            'depositario': r['depositario'],
            'grupo': r['grupo'],
            'category': fee_data.get('category', ''),
            'depo_fee': fee_data.get('depositary_fee'),
            'patrimonio_k': fee_data.get('patrimonio_k'),
            'mgmt_fee': fee_data.get('mgmt_fee_aum'),
            'ter': fee_data.get('ter'),
            'investors': fee_data.get('investors', 0),
            'fund_type': 'FI',
        })

    # Append SICAV series records (depositary fees aren't published for SICAVs in
    # the public CNMV feed, so depo_fee/mgmt_fee/ter are left None — they're handled
    # downstream by the AUM-weighted aggregations which already skip None entries).
    sicav_aum_k = 0.0
    for r in sicav_records:
        sicav_aum_k += r['patrimonio_k']
        joined.append({
            'fund_name': r['fund_name'],
            'isin': r['isin'],
            'share_class': r['share_class'],
            'gestora': r['gestora'],
            'depositario': r['depositario'],
            'grupo': r['grupo'],
            'category': r['category'] or 'SICAV',
            'depo_fee': None,
            'patrimonio_k': r['patrimonio_k'],
            'mgmt_fee': None,
            'ter': None,
            'investors': r['investors'],
            'fund_type': 'SICAV',
        })

    # === Fund-level aggregation (sum AUM across all classes per fund) ===
    _fund_classes = {}
    for r in joined:
        key = r['fund_name']
        if key not in _fund_classes:
            _fund_classes[key] = []
        _fund_classes[key].append(r)

    fund_level = {}
    for fname, classes in _fund_classes.items():
        base = dict(classes[0])
        total_patrim = sum(c['patrimonio_k'] or 0 for c in classes)
        base['patrimonio_k'] = total_patrim
        base['investors'] = sum(c.get('investors') or 0 for c in classes)
        # AUM-weighted depo fee
        dep_num = sum((c['depo_fee'] or 0) * (c['patrimonio_k'] or 0)
                      for c in classes if c['depo_fee'] is not None and c['depo_fee'] > 0)
        dep_den = sum(c['patrimonio_k'] or 0
                      for c in classes if c['depo_fee'] is not None and c['depo_fee'] > 0)
        base['depo_fee'] = round(dep_num / dep_den, 4) if dep_den > 0 else base['depo_fee']
        # AUM-weighted TER
        ter_num = sum((c['ter'] or 0) * (c['patrimonio_k'] or 0)
                      for c in classes if c['ter'] is not None and c['ter'] > 0)
        ter_den = sum(c['patrimonio_k'] or 0
                      for c in classes if c['ter'] is not None and c['ter'] > 0)
        base['ter'] = round(ter_num / ter_den, 4) if ter_den > 0 else base['ter']
        fund_level[fname] = base

    funds = list(fund_level.values())
    total_aum_k = sum(f['patrimonio_k'] or 0 for f in funds)

    # === Depositario Stats ===
    depo_groups = {}
    for f in funds:
        d = f['depositario']
        if d not in depo_groups:
            depo_groups[d] = []
        depo_groups[d].append(f)

    depositario_stats = []
    for d, d_funds in sorted(depo_groups.items(), key=lambda x: -sum(f['patrimonio_k'] or 0 for f in x[1])):
        aum_k = sum(f['patrimonio_k'] or 0 for f in d_funds)
        depo_fees = [f['depo_fee'] for f in d_funds if f['depo_fee'] is not None and f['depo_fee'] > 0]
        gestoras = set(f['gestora'] for f in d_funds)

        # AUM and counts split by fund_type (FI vs SICAV — alternativos appended below)
        fi_funds    = [f for f in d_funds if f.get('fund_type') == 'FI']
        sicav_funds = [f for f in d_funds if f.get('fund_type') == 'SICAV']
        fi_aum_k    = sum(f['patrimonio_k'] or 0 for f in fi_funds)
        sicav_aum_kk = sum(f['patrimonio_k'] or 0 for f in sicav_funds)

        depositario_stats.append({
            'depositario': d,
            'gestora_count': len(gestoras),
            'fund_count': len(d_funds),
            'total_aum_bn': round(aum_k / 1_000_000, 1),
            'market_share_pct': round(aum_k / total_aum_k * 100, 1) if total_aum_k > 0 else 0,
            'avg_depo_fee': round(sum(depo_fees) / len(depo_fees), 4) if depo_fees else 0,
            'median_depo_fee': _median(depo_fees),
            'is_inversis': INVERSIS_NAME in d,
            # Per-asset-type breakdown (FI + SICAV from CNMV bulk feeds; alternativos
            # added in a later loop once the alternativos scrape is processed below).
            'fi_aum_bn':    round(fi_aum_k / 1_000_000, 2),
            'fi_fund_count': len(fi_funds),
            'sicav_aum_bn': round(sicav_aum_kk / 1_000_000, 2),
            'sicav_fund_count': len(sicav_funds),
            'fil_est_aum_bn': 0.0,   # filled in below from alternativos scrape
            'fil_entity_count': 0,
            'alt_entity_count': 0,
            # Published overlay placeholders — set later from Funds People file.
            'pension_aum_bn': None,
            'epsv_aum_bn': None,
            'alt_published_aum_bn': None,
            'published_total_aum_bn': None,
            'published_source': None,
            'combined_aum_bn': round(aum_k / 1_000_000, 2),  # recomputed below
        })

    # === Gestora-Depositario relationships ===
    gd_key = {}  # (gestora, depositario) -> list of funds
    for f in funds:
        key = (f['gestora'], f['depositario'])
        if key not in gd_key:
            gd_key[key] = []
        gd_key[key].append(f)

    # Inversis avg fee for opportunity comparison
    inversis_funds = [f for f in funds if INVERSIS_NAME in f['depositario']]
    inversis_fees = [f['depo_fee'] for f in inversis_funds if f['depo_fee'] is not None and f['depo_fee'] > 0]
    inversis_avg_fee = round(sum(inversis_fees) / len(inversis_fees), 4) if inversis_fees else 0.075

    gestora_depositario = []
    for (gestora, depositario), gd_funds in sorted(gd_key.items(), key=lambda x: -sum(f['patrimonio_k'] or 0 for f in x[1])):
        aum_k = sum(f['patrimonio_k'] or 0 for f in gd_funds)
        depo_fees = [f['depo_fee'] for f in gd_funds if f['depo_fee'] is not None and f['depo_fee'] > 0]
        grupo = gd_funds[0]['grupo']

        # Weighted avg depo fee by AUM
        weighted_num = sum((f['depo_fee'] or 0) * (f['patrimonio_k'] or 0)
                          for f in gd_funds if f['depo_fee'] is not None and f['depo_fee'] > 0)
        weighted_den = sum(f['patrimonio_k'] or 0
                          for f in gd_funds if f['depo_fee'] is not None and f['depo_fee'] > 0)
        weighted_fee = round(weighted_num / weighted_den, 4) if weighted_den > 0 else 0

        # Category mix
        cat_map = {}
        for f in gd_funds:
            cat = f['category'] or 'Unknown'
            if cat not in cat_map:
                cat_map[cat] = {'aum_k': 0, 'funds': 0}
            cat_map[cat]['aum_k'] += f['patrimonio_k'] or 0
            cat_map[cat]['funds'] += 1
        category_mix = sorted([
            {'category': cat, 'aum_m': round(v['aum_k'] / 1000, 1), 'funds': v['funds']}
            for cat, v in cat_map.items()
        ], key=lambda x: -x['aum_m'])

        gestora_depositario.append({
            'gestora': gestora,
            'depositario': depositario,
            'grupo': grupo,
            'fund_count': len(gd_funds),
            'total_aum_m': round(aum_k / 1000, 1),
            'avg_depo_fee': round(sum(depo_fees) / len(depo_fees), 4) if depo_fees else 0,
            'weighted_depo_fee': weighted_fee,
            'is_captive': _is_captive(grupo, depositario),
            'is_inversis': INVERSIS_NAME in depositario,
            'category_mix': category_mix[:5],  # top 5 categories
        })

    # === Fund Detail ===
    fund_detail = []
    for f in funds:
        fund_detail.append({
            'fund_name': f['fund_name'],
            'isin': f['isin'],
            'gestora': f['gestora'],
            'depositario': f['depositario'],
            'category': f['category'] or 'Unknown',
            'aum_m': round((f['patrimonio_k'] or 0) / 1000, 1),
            'depo_fee': f['depo_fee'],
            'mgmt_fee': f['mgmt_fee'],
            'ter': f['ter'],
            'is_inversis': INVERSIS_NAME in f['depositario'],
        })

    # === Opportunity Targets ===
    # Group by gestora (across all their depositarios)
    gestora_all = {}
    for f in funds:
        g = f['gestora']
        if g not in gestora_all:
            gestora_all[g] = []
        gestora_all[g].append(f)

    opportunity_targets = []
    for gestora, g_funds in gestora_all.items():
        # Skip if already Inversis client
        depos = set(f['depositario'] for f in g_funds)
        if INVERSIS_NAME in depos:
            continue

        grupo = g_funds[0]['grupo']
        primary_depo = max(depos, key=lambda d: sum(
            f['patrimonio_k'] or 0 for f in g_funds if f['depositario'] == d))

        is_captive = _is_captive(grupo, primary_depo)

        aum_k = sum(f['patrimonio_k'] or 0 for f in g_funds)
        depo_fees = [f['depo_fee'] for f in g_funds if f['depo_fee'] is not None and f['depo_fee'] > 0]
        avg_fee = round(sum(depo_fees) / len(depo_fees), 4) if depo_fees else 0

        # Category mix
        cat_map = {}
        for f in g_funds:
            cat = f['category'] or 'Unknown'
            if cat not in cat_map:
                cat_map[cat] = {'aum_k': 0, 'funds': 0}
            cat_map[cat]['aum_k'] += f['patrimonio_k'] or 0
            cat_map[cat]['funds'] += 1
        category_mix = sorted([
            {'category': cat, 'aum_m': round(v['aum_k'] / 1000, 1), 'funds': v['funds']}
            for cat, v in cat_map.items()
        ], key=lambda x: -x['aum_m'])

        opportunity_targets.append({
            'gestora': gestora,
            'current_depositario': primary_depo,
            'grupo': grupo,
            'total_aum_m': round(aum_k / 1000, 1),
            'fund_count': len(g_funds),
            'avg_depo_fee': avg_fee,
            'fee_vs_inversis_avg': round(avg_fee - inversis_avg_fee, 4),
            'potential_revenue_k': round(aum_k * avg_fee / 100, 0) if avg_fee > 0 else 0,
            'is_captive': is_captive,
            'category_mix': category_mix[:3],
        })

    # Sort: non-captive first, then by AUM
    opportunity_targets.sort(key=lambda x: (x['is_captive'], -x['total_aum_m']))

    # === Market Share Chart (short names for display) ===
    market_share_chart = []
    for ds in depositario_stats:
        short = ds['depositario'].replace(', S.A.', '').replace(', S.A', '').replace(' S.A.', '')
        market_share_chart.append({
            'depositario': short,
            'depositario_full': ds['depositario'],
            'aum_bn': ds['total_aum_bn'],
            'pct': ds['market_share_pct'],
            'is_inversis': ds['is_inversis'],
        })

    # === Fee by Depositario (for bar chart) ===
    fee_by_depositario = []
    for ds in depositario_stats:
        short = ds['depositario'].replace(', S.A.', '').replace(', S.A', '').replace(' S.A.', '')
        fee_by_depositario.append({
            'depositario': short,
            'depositario_full': ds['depositario'],
            'avg_fee': ds['avg_depo_fee'],
            'median_fee': ds['median_depo_fee'],
            'aum_bn': ds['total_aum_bn'],
            'is_inversis': ds['is_inversis'],
        })

    # === Summary ===
    inversis_stats = next((d for d in depositario_stats if d['is_inversis']), None)
    inversis_rank_aum = next((i + 1 for i, d in enumerate(depositario_stats) if d['is_inversis']), 0)
    # Rank by gestora count
    by_gestora_count = sorted(depositario_stats, key=lambda x: -x['gestora_count'])
    inversis_rank_gestoras = next((i + 1 for i, d in enumerate(by_gestora_count) if d['is_inversis']), 0)

    # Opportunity summary
    non_captive_targets = [t for t in opportunity_targets if not t['is_captive']]
    premium_targets = [t for t in non_captive_targets if t['fee_vs_inversis_avg'] > 0]

    summary = {
        'total_depositarios': len(depositario_stats),
        'total_gestoras': len(gestora_all),
        'total_funds': len(funds),
        'total_aum_bn': round(total_aum_k / 1_000_000, 1),
        'inversis_aum_bn': inversis_stats['total_aum_bn'] if inversis_stats else 0,
        'inversis_market_share_pct': inversis_stats['market_share_pct'] if inversis_stats else 0,
        'inversis_gestora_count': inversis_stats['gestora_count'] if inversis_stats else 0,
        'inversis_fund_count': inversis_stats['fund_count'] if inversis_stats else 0,
        'inversis_avg_fee': inversis_avg_fee,
        'inversis_rank_aum': inversis_rank_aum,
        'inversis_rank_gestoras': inversis_rank_gestoras,
        'addressable_aum_bn': round(sum(t['total_aum_m'] for t in non_captive_targets) / 1000, 1),
        'target_gestoras': len(premium_targets),
        'potential_revenue_m': round(sum(t['potential_revenue_k'] for t in premium_targets) / 1000, 1),
        'date': cnmv_date,
        'sicav_period': sicav_period,
        'sicav_fund_count': sum(1 for r in joined if r.get('fund_type') == 'SICAV'),
        'sicav_aum_bn': round(sicav_aum_k / 1_000_000, 2),
    }

    all_depositarios = sorted(set(f['depositario'] for f in funds))
    all_gestoras = sorted(set(f['gestora'] for f in funds))
    all_categories = sorted(set(f['category'] for f in funds if f['category']))

    # ── Enrich summary + charts from inversis_depositary_insights ──────────
    # Use the depositary_ranking sheet for true market-wide AUM figures
    rank_records = insights.get('depositary_ranking', [])
    if rank_records:
        total_market_aum_k = sum(_sf(r.get('aum_k')) for r in rank_records)
        mkt_chart = []
        for r in sorted(rank_records, key=lambda x: -_sf(x.get('aum_k'))):
            depo = str(r.get('Entidad Depositaria', ''))
            short = depo.replace(', S.A.', '').replace(', S.A', '').replace(' S.A.', '')
            aum_k = _sf(r.get('aum_k'))
            mkt_chart.append({
                'depositario': short[:40],
                'depositario_full': depo,
                'aum_bn': round(aum_k / 1_000_000, 1),
                'funds': int(_sf(r.get('funds', 0))),
                'pct': round(aum_k / total_market_aum_k * 100, 2) if total_market_aum_k else 0,
                'is_inversis': 'INVERSIS' in depo,
            })
        summary['market_total_aum_bn'] = round(total_market_aum_k / 1_000_000, 1)
    else:
        mkt_chart = market_share_chart  # fallback to CNMV-derived chart

    # Inversis gestora-level data from insights (richer than CNMV Anexo alone)
    inv_by_gestora = insights.get('inversis_by_gestora', [])
    total_rev_k = sum(_sf(r.get('est_deposit_fee_rev_k')) for r in inv_by_gestora)
    summary['inversis_est_annual_rev_m'] = round(total_rev_k / 1000, 2)

    inv_by_gestora_out = []
    for r in inv_by_gestora:
        g = str(r.get('Sociedad Gestora', ''))
        inv_by_gestora_out.append({
            'gestora': g,
            'gestora_short': g.split(',')[0][:40],
            'classes': int(_sf(r.get('classes', 0))),
            'funds': int(_sf(r.get('funds', 0))),
            'aum_m': round(_sf(r.get('aum_k')) / 1000, 1),
            'avg_fee_bps': round(_sf(r.get('avg_deposit_fee_pct')) * 100, 2),
            'weighted_fee_bps': round(_sf(r.get('aum_w_deposit_fee_pct')) * 100, 2),
            'est_annual_rev_k': round(_sf(r.get('est_deposit_fee_rev_k')), 1),
            'is_march': 'MARCH ASSET' in g,
        })

    # ── Alternativos: aggregate entity counts per depositary ──────────────
    # AUM is not available per entity (CNMV public pages don't expose it).
    # We expose entity counts and a per-depositary FIL AUM estimate (proportional
    # split of the €8.0B Cuadro 6.1 aggregate by FIL entity count).
    FIL_AGGREGATE_AUM_K = 7_989_231  # IICIL €7,156,484k + IICIICIL €832,747k (Q2 2025)
    fil_total_entities = sum(1 for r in alternativos_records if r['fund_type'].startswith('FIL'))

    alt_by_depo = {}
    for r in alternativos_records:
        d = r['depositario']
        if d not in alt_by_depo:
            alt_by_depo[d] = {'total': 0, 'by_type': {}, 'fil_count': 0}
        alt_by_depo[d]['total'] += 1
        ftype = r['fund_type']
        alt_by_depo[d]['by_type'][ftype] = alt_by_depo[d]['by_type'].get(ftype, 0) + 1
        if ftype.startswith('FIL'):
            alt_by_depo[d]['fil_count'] += 1

    alternativos_stats = []
    for d, info in sorted(alt_by_depo.items(), key=lambda x: -x[1]['total']):
        fil_aum_est_k = (info['fil_count'] / fil_total_entities * FIL_AGGREGATE_AUM_K
                         if fil_total_entities else 0)
        alternativos_stats.append({
            'depositario': d,
            'entity_count': info['total'],
            'fil_count': info['fil_count'],
            'fil_aum_est_bn': round(fil_aum_est_k / 1_000_000, 2),
            'by_type': info['by_type'],
            'is_inversis': INVERSIS_NAME in d,
        })

    inversis_alt = next((a for a in alternativos_stats if a['is_inversis']), None)
    inversis_alt_rank = next(
        (i + 1 for i, a in enumerate(alternativos_stats) if a['is_inversis']), 0)

    # Enrich depositario_stats with alternativos counts + FIL est AUM, and recompute
    # combined_aum_bn = FI + SICAV + (published alts || FIL est) + (published pensiones)
    # + (published EPSV). Some depositarios appear only in alternativos (no FI/SICAV) —
    # append those as new entries with zero FI/SICAV.
    alt_lookup = {a['depositario']: a for a in alternativos_stats}
    existing_depos = {ds['depositario'] for ds in depositario_stats}
    published = _load_published_overlay()
    for ds in depositario_stats:
        a = alt_lookup.get(ds['depositario'])
        if a:
            ds['fil_est_aum_bn']    = a['fil_aum_est_bn']
            ds['fil_entity_count']  = a['fil_count']
            ds['alt_entity_count']  = a['entity_count']
        # Apply Funds People overlay (pensions, EPSV, published alts; optional FI/SICAV)
        pub = _match_published(ds['depositario'], published)
        ds['fi_aum_bn_cnmv']    = ds['fi_aum_bn']     # preserve CNMV-derived for diff
        ds['sicav_aum_bn_cnmv'] = ds['sicav_aum_bn']
        if pub:
            if pub.get('fi') is not None:
                ds['fi_aum_bn'] = round(pub['fi'] / 1000, 3)
            if pub.get('sicav') is not None:
                ds['sicav_aum_bn'] = round(pub['sicav'] / 1000, 3)
            if pub.get('pensiones') is not None:
                ds['pension_aum_bn'] = round(pub['pensiones'] / 1000, 3)
            if pub.get('epsv') is not None:
                ds['epsv_aum_bn'] = round(pub['epsv'] / 1000, 3)
            if pub.get('alternativos') is not None:
                ds['alt_published_aum_bn'] = round(pub['alternativos'] / 1000, 3)
            if pub.get('total') is not None:
                ds['published_total_aum_bn'] = round(pub['total'] / 1000, 3)
            ds['published_source'] = pub.get('source') or pub.get('source_other')
        # Use published alts if available, else FIL-only proportional estimate
        alt_for_combined = (ds['alt_published_aum_bn']
                            if ds['alt_published_aum_bn'] is not None
                            else ds['fil_est_aum_bn'])
        pension_for_combined = ds['pension_aum_bn'] or 0
        epsv_for_combined    = ds['epsv_aum_bn'] or 0
        ds['combined_aum_bn'] = round(
            ds['fi_aum_bn'] + ds['sicav_aum_bn'] + alt_for_combined
            + pension_for_combined + epsv_for_combined, 2)
    # Append alternativos-only depositarios (none in current data, but defensive)
    for a in alternativos_stats:
        if a['depositario'] in existing_depos:
            continue
        depositario_stats.append({
            'depositario': a['depositario'],
            'gestora_count': 0,
            'fund_count': 0,
            'total_aum_bn': 0,
            'market_share_pct': 0,
            'avg_depo_fee': 0,
            'median_depo_fee': 0,
            'is_inversis': a['is_inversis'],
            'fi_aum_bn': 0.0,
            'fi_fund_count': 0,
            'sicav_aum_bn': 0.0,
            'sicav_fund_count': 0,
            'fil_est_aum_bn':    a['fil_aum_est_bn'],
            'fil_entity_count':  a['fil_count'],
            'alt_entity_count':  a['entity_count'],
            'pension_aum_bn': None,
            'epsv_aum_bn': None,
            'alt_published_aum_bn': None,
            'published_total_aum_bn': None,
            'published_source': None,
            'combined_aum_bn':   round(a['fil_aum_est_bn'], 2),
        })
    # Re-sort by combined AUM (alternativos + pensions can shift ranking)
    depositario_stats.sort(key=lambda x: -x['combined_aum_bn'])

    # Inversis combined position (FI + SICAV + FIL est)
    inversis_combined = next((d for d in depositario_stats if d['is_inversis']), None)
    inversis_combined_rank = next(
        (i + 1 for i, d in enumerate(depositario_stats) if d['is_inversis']), 0)

    summary['alternativos_total_entities'] = len(alternativos_records)
    summary['alternativos_total_depositarios'] = len(alternativos_stats)
    summary['alternativos_fil_aum_bn'] = round(FIL_AGGREGATE_AUM_K / 1_000_000, 1)
    summary['inversis_alt_entities'] = inversis_alt['entity_count'] if inversis_alt else 0
    summary['inversis_alt_fil_count'] = inversis_alt['fil_count'] if inversis_alt else 0
    summary['inversis_alt_fil_aum_bn'] = inversis_alt['fil_aum_est_bn'] if inversis_alt else 0
    summary['inversis_alt_rank'] = inversis_alt_rank
    # Combined headline: FI + SICAV + best-available alts + pensions + EPSV
    summary['inversis_combined_aum_bn'] = inversis_combined['combined_aum_bn'] if inversis_combined else 0
    summary['inversis_combined_rank']   = inversis_combined_rank
    summary['combined_market_aum_bn']   = round(
        sum(d['combined_aum_bn'] for d in depositario_stats), 1)
    summary['inversis_combined_share_pct'] = round(
        inversis_combined['combined_aum_bn'] / summary['combined_market_aum_bn'] * 100, 2
    ) if inversis_combined and summary['combined_market_aum_bn'] else 0

    # ── Published Funds People IX overlay summary ─────────────────────────
    if published:
        mkt_totals = published.get('market_totals_2025Q4', {})
        summary['published_total_aum_bn']        = round(
            (mkt_totals.get('fundspeople_total_aum_m') or 0) / 1000, 1)
        summary['published_pensions_total_bn']   = round(
            (mkt_totals.get('pensiones_aum_m') or 0) / 1000, 1)
        summary['published_pensions_individual_bn'] = round(
            (mkt_totals.get('pensiones_individual_m') or 0) / 1000, 1)
        summary['published_pensions_empleo_bn']  = round(
            (mkt_totals.get('pensiones_empleo_m') or 0) / 1000, 1)
        summary['published_epsv_total_bn']       = round(
            (mkt_totals.get('epsv_aum_m_estimate') or 0) / 1000, 1)
        summary['published_alternativos_total_bn'] = round(
            (mkt_totals.get('alternativos_aum_m_estimate') or 0) / 1000, 1)
        summary['published_source'] = published['_metadata']['primary_source']

        # Inversis published breakdown (verbatim from FundsPeople IX 2026)
        inv_pub = next((r for r in published['depositaries']
                        if 'INVERSIS' in (r.get('depositario_match') or '').upper()), None)
        if inv_pub:
            summary['inversis_published'] = {
                'fi_bn':           round((inv_pub.get('fi') or 0) / 1000, 3),
                'sicav_bn':        round((inv_pub.get('sicav') or 0) / 1000, 3),
                'pensiones_bn':    round((inv_pub.get('pensiones') or 0) / 1000, 3),
                'epsv_bn':         round((inv_pub.get('epsv') or 0) / 1000, 3),
                'alternativos_bn': round((inv_pub.get('alternativos') or 0) / 1000, 3),
                'total_bn':        round((inv_pub.get('total') or 0) / 1000, 3),
                'source':          inv_pub.get('source'),
            }

        bnp_pub = next((r for r in published['depositaries']
                        if 'BNP PARIBAS' in (r.get('depositario_match') or '').upper()), None)
        if bnp_pub and bnp_pub.get('alternativos') is not None:
            summary['bnp_published_alts_bn'] = round(bnp_pub['alternativos'] / 1000, 2)

    # Qualitative analysis section
    qualitative_analysis = _build_qualitative_analysis(
        insights, summary, depositario_stats, opportunity_targets)

    return {
        'summary': summary,
        'depositario_stats': depositario_stats,
        'gestora_depositario': gestora_depositario,
        'fund_detail': fund_detail,
        'opportunity_targets': opportunity_targets,
        'market_share_chart': mkt_chart,            # now from insights depositary_ranking
        'fee_by_depositario': fee_by_depositario,
        'inversis_by_gestora': inv_by_gestora_out,  # NEW: Inversis client detail with revenue
        'qualitative_analysis': qualitative_analysis,  # NEW: computed qualitative insights
        'alternativos_stats': alternativos_stats,   # NEW: FIL/FCR/SCR/SICC/FICC entity counts
        'depositarios': all_depositarios,
        'gestoras': all_gestoras,
        'categories': all_categories,
    }
